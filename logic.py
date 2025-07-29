from openpyxl import load_workbook

file_path = './NEW CODE & REPAIR.xlsx'
sheet_name = 'Sheet1'

# Define the Part families
part_families = {
    'No Power': ['L1'],
    'No Display': ['L2'],
    'Hanging': ['L3'],
    'Auto Restart': ['L4'],
    'Power Button Not Working': ['L5'],
    'BIOS Password': ['L6'],
    'Power On Password': ['L7'],
    'HDD Password': ['L8'],
    'Body': ['L9', 'L10', 'L11', 'L12', 'L13', 'L14', 'L15', 'L16', 'L17', 'L18', 'L19', 'L20'],
    'Screen': ['L21', 'L22', 'L23', 'L24', 'L25', 'L26', 'L27', 'L28', 'L29', 'L30', 'L31'],
    'Hinges': ['L32', 'L33', 'L34'],
    'HDD': ['L35', 'L36', 'L37'],
    'Memory Missing': ['L38'],
    'Keys': ['L39', 'L40', 'L41'],
    'Touchpad Not Working': ['L42'],
    'Camera Not Working': ['L43'],
    'Speaker': ['L44', 'L45'],
    'WIFI Not Working': ['L46'],
    'Bluetooth Not Working': ['L47'],
    'Mic Not Working': ['L48'],
    'USB Not Working': ['L49'],
    'Battery': ['L50', 'L51', 'L52'],
    'Charger': ['L53', 'L54', 'L55', 'L56'],
    'Carry Bag Missing': ['L57'],
    'Carry Bag Damaged': ['L58']
}


def load_file_and_sheet(file_path, sheet_name=None):
    """    
    Loads the Excel file and sheet
    Args:
        file_path (str): Path to the Excel file
        sheet_name (str, optional): Name of the sheet. If None, uses active sheet
    """
    global wb, ws

    try:
        # Load the workbook
        wb = load_workbook(file_path)
        
        # Select the worksheet
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        # print(f'Successfully Loaded {sheet_name} from {file_path}') 
    except FileNotFoundError:
        # print(f"Error: File '{file_path}' not found.")
        return None
    except Exception as e:
        # print(f"Error: {e}")
        return None
    
def search_and_get_single_cost(search_value):
    """
    Search for a value in column 1 and return the corresponding value from column 3
    
    Args:
        search_value: Value to search for in column 1
    Returns:
        The value from column 3 of the matching row, or None if not found
    """

    # Search through column 1
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        
        # Check if the value matches (case-insensitive for strings)
        if cell_value == search_value or (
            isinstance(cell_value, str) and isinstance(search_value, str) and 
            cell_value.lower() == search_value.lower()
        ):
            # Return the value from column 3 of the same row
            single_cost = ws.cell(row=row, column=3).value

            if single_cost == 'NTR':
                single_cost = 'NEED TO BE REPLACED'
            elif single_cost == 'NA':
                single_cost = 'NOT AVAILABLE'
            else:
                single_cost = int(single_cost)

            # print(f"Found '{search_value}' in column 1.")
            # print(f"Corresponding value in column 3 ( single cost ): {single_cost}")

            return single_cost

    # If no match found
    print(f"'{search_value}' not found in column 1.")
    return None
        
def search_and_get_max_cost(search_value):
    """
    Search for a value in column 1 and return the corresponding value from column 4
    Also checks whether the cells in column 4 are part of a merged range
    
    Args:
        search_value: Value to search for in column 1
    Returns:
        The value from column 4 of the matching row, or None if not found
    """

    # Search through column 1
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        
        # Check if the value matches (case-insensitive for strings)
        if cell_value == search_value or (
            isinstance(cell_value, str) and isinstance(search_value, str) and 
            cell_value.lower() == search_value.lower()
        ):
            # Return the value from column 4 of the same row
            column4_cell = ws.cell(row=row, column=4)
            max_cost = column4_cell.value

            if max_cost is None:
                # Check if this cell is part of a merged range
                for merged_range in ws.merged_cells.ranges:
                    if column4_cell.coordinate in merged_range:
                        # Get the top-left cell of the merged range (which contains the value)
                        top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        max_cost = top_left_cell.value

            if max_cost is not None:
                max_cost = int(max_cost)

            # print(f"Found '{search_value}' in column 1.")
            # print(f"Corresponding value in column 4 ( max cost of the corresponding part ): {max_cost}")

            return max_cost
    
    # If no match found
    print(f"'{search_value}' not found in column 1.")
    return None

def find_family(codes):
    """
    Finds which of the given codes are related to the same family ( same part )

    Args:
        code1, code2, code3, code4
    Returns:
        part_family_groups

    Called by 'calculate_cost()' function
    """

    # Find which family each code belongs to
    code_families = {}  # Dictionary with items in the format of {'code': 'part-family'}
    for code in codes:
        for family_name, family_list in part_families.items():
            if code in family_list:
                code_families[code] = family_name
                break

    # # Display results
    # print("Code Family Analysis:")
    # print("=" * 30)

    # for code in codes:
    #     if code in code_families:
    #         print(f"{code} belongs to: {code_families[code]}")
    #     else:
    #         print(f"{code} belongs to: No family found")

    # print("\n" + "=" * 30)

    # Check if codes are in the same Part family
    part_family_groups = {} # Dictionary with items in the format of {'part-family': ['code1', 'code2']} if code1, code2 belong to part-family
    for code, family in code_families.items():
        if family not in part_family_groups:
            part_family_groups[family] = []
        part_family_groups[family].append(code)

    # print("Codes grouped by family:")
    # for family, codes_in_family in part_family_groups.items():
        # print(f"Family '{family}': {codes_in_family}")

    return part_family_groups

def calculate_costs_of_individual_families(family_array):
    """
    Calculates cost of Individual Part Families 

    Args:
        family_array - Array that contains Codes that are present in the Particular Family
    Returns:
        cost_of_family

    Called by main 'calculate_cost()' function
    """
    # print(family_array)

    cost_of_family = 0

    for code in family_array:
        part_cost = search_and_get_single_cost(code);

        if type(part_cost) == int:
            cost_of_family += part_cost
        elif type(part_cost) == str: # part_cost = "NEED TO BE REPLACED" or "NOT AVAILABLE"
            cost_of_family = part_cost

    max_cost_of_family = search_and_get_max_cost(family_array[0])

    if max_cost_of_family is not None and cost_of_family > max_cost_of_family:
        cost_of_family = max_cost_of_family

    return cost_of_family

def calculate_cost(codes=[None, None, None, None]):
    """
    Main Function that calculates the final costs
    """

    to_be_replaced = False
    total_cost = 0

    load_file_and_sheet(file_path, sheet_name)

    part_family_groups = find_family(codes)

    if part_family_groups: # If Row is Empty, part_family_groups will be an empty dictionary
        pass
        # print("Cost Breakdown:")
    
    for name_of_family, codes_in_family in part_family_groups.items():
        cost_of_family = calculate_costs_of_individual_families(codes_in_family)

        if cost_of_family == "NEED TO BE REPLACED":
            to_be_replaced = True
            total_cost = cost_of_family

    for name_of_family, codes_in_family in part_family_groups.items():
        if not to_be_replaced:
            cost_of_family = calculate_costs_of_individual_families(codes_in_family)
            # print(f"{name_of_family} : {cost_of_family}")
            total_cost = total_cost + cost_of_family

    # print(f"Total cost : {total_cost}")
    return total_cost
    wb.close()

if __name__ == "__main__":
    calculate_cost()
