from openpyxl import load_workbook
import sys
import logic

header_row = 2

def load_input_file_and_sheet(input_file_path, input_sheet_name="INSP SHEET"):
    """    
    Loads the Excel file and sheet
    Args:
        file_path (str): Path to the Excel file
        sheet_name (str, optional): Name of the sheet. If None, uses active sheet
    """
    global workbook, worksheet

    try:
        # Load the workbook
        workbook = load_workbook(input_file_path)
        
        # Select the worksheet
        if input_sheet_name:
            worksheet = workbook[input_sheet_name]
        else:
            worksheet = workbook.active
        # print(f'Successfully Loaded {input_sheet_name} from {input_file_path}')

    except FileNotFoundError:
        # print(f"Error: File '{file_path}' not found.")
        return None
    except Exception as e:
        print(f"Error: {e}")
        input_sheet_name = input("Enter Sheet Name which contains Inspection Data (for ex: Sheet1/Sheet2/Sheet3): ")
        load_input_file_and_sheet(input_file_path, input_sheet_name)
        return None
    
def get_code_values_for_row(row_number):
    """
    Extract values from FUNCTIONING, BODY-A, BODY-B, BODY-C, BODY-D, SCREEN-A, SCREEN-B, SCREEN-C, N/W-A, N/W-B, MISSING-A, MISSING-B columns for a specific row
    
    Args:
        row_number (int): Row number to extract data from (2-indexed)
    
    Returns:
        code_values: Array containing the code values
    """


    target_columns = ['FUNCTIONING' ,'BODY-A', 'BODY-B', 'BODY-C', 'BODY-D', 'SCREEN-A', 'SCREEN-B', 'SCREEN-C', 'N/W-A', 'N/W-B', 'MISSING-A', 'MISSING-B']
    
    # Find the column indices for the target columns
    column_indices = {}
    # column_indices = {'FUNCTIONING': 39,'BODY-A': 40, 'BODY-B': 41, 'BODY-C': 42, 'BODY-D': 43, 'SCREEN-A': 44, 'SCREEN-B':45, 'SCREEN-C':46, 'N/W-A':47, '4N/W-B':48, 'MISSING-A':49, 'MISSING-B':50}
    
    
    # Iterate through the header_row to find column headers
    for col in range(1, worksheet.max_column + 1):
        header_value = worksheet.cell(row=header_row, column=col).value
        if header_value in target_columns:
            # print(f"{header_value} found at {col}")
            target_columns.remove(header_value)
            column_indices[header_value] = col
    
    # Check if all required columns were found
    missing_columns = set(target_columns) - set(column_indices.keys())
    if missing_columns:
        print(f"Warning: Missing columns: {missing_columns}")
    
    # Extract values for the specified row
    code_values = []

    # for column_name in target_columns:
    #     if column_name in column_indices:
    #         col_index = column_indices[column_name]
    #         cell_value = worksheet.cell(row=row_number, column=col_index).value
    #         print(f'{row_number}, {col_index} : {cell_value}')
    #         code_values.append(cell_value)
    #     else:
    #         code_values.append(None)  # Add None for missing columns

    for column_name in column_indices.keys():
        col_index = column_indices[column_name]
        cell_value = worksheet.cell(row=row_number, column=col_index).value
        # print(f'{row_number}, {col_index} : {cell_value}')
        code_values.append(cell_value)

    return code_values

def check_and_create_total_cost_column():
    """
    Check if 'Total Cost' column exists, create it if it doesn't.
    
    Returns: 
        total_cost_column : The column number of the Total Cost Column
    """

    # Check if 'Total Cost' column exists in the first row (headers)
    total_cost_exists = False
    total_cost_column = None
    
    # Iterate through the header_row to find headers
    for cell in worksheet[header_row]:  # Header row
        if cell.value and str(cell.value).strip().lower() == "total cost":
            total_cost_exists = True
            total_cost_column = cell.column
            break
    
    if total_cost_exists:
        pass
        # print(f"'Total Cost' column already exists at column {total_cost_column}")
    else:
        # Find the next available column
        max_col = worksheet.max_column
        next_col = max_col + 1
        
        # Add 'Total Cost' header to the next available column
        worksheet.cell(row=header_row, column=next_col, value="Total Cost")
        total_cost_column = next_col
        print(f"'Total Cost' column created at column {next_col}")
    
    return total_cost_column

def add_row_cost_to_sheet(row, column, row_cost):
    """
    Adds the Total cost of the Row to the Total Cost Column in the Excel Sheet.

    Args:
        row 
        column
        row_cost
    """

    worksheet.cell(row=row, column=column, value=row_cost)
    # print(f"{row_cost} inserted at Row: {row} Column: {column}")

def main():
    """
    Main Function

    Iterates over each row and Calls 'get_code_values_for_row' and 'backend.calculate_cost' functions for all rows
    """

    for row in range(header_row+1, worksheet.max_row + 1): # Start From header_row + 1 because header_row contains only Headers
        codes = get_code_values_for_row(row_number=row)

        all_null = all(element is None for element in codes) # True if code has all None values

        if not all_null:
            # print(f'Row {row}')
            # print(codes)

            total_cost = logic.calculate_cost(codes)

            # print(total_cost)

            column = check_and_create_total_cost_column()
            add_row_cost_to_sheet(row, column, total_cost)

            # print("\n")

def run(input_file_path):
    try:
        load_input_file_and_sheet(input_file_path)

        main()
        workbook.save(input_file_path)
        workbook.close()
    except Exception as e:
        print("File Not Found. Please make sure the filename is correct")


if len(sys.argv) == 1:
    # input_file_path = input("Enter Input Filename: \n")
    from tkinter import Tk
    from tkinter.filedialog import askopenfilename

    Tk().withdraw()
    input_file_path = askopenfilename(title="Select One or More files", defaultextension=".xlsx") 
    run(input_file_path)

else:
    # Take file name from command line argument
    for i in range(1, len(sys.argv)):
        # print(sys.argv[i], end = " ")

        input_file_path = sys.argv[i]
        run(input_file_path)
